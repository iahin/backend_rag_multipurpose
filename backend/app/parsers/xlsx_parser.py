from io import BytesIO

from openpyxl import load_workbook

from app.models.schemas import NormalizedDocument, ParsedFile
from app.parsers.base import FileParser


class XlsxParser(FileParser):
    supported_types = ("xlsx",)

    async def parse(
        self,
        filename: str,
        content: bytes,
        mime_type: str | None,
        source_type_override: str | None = None,
        shared_metadata: dict | None = None,
        tags: list[str] | None = None,
    ) -> ParsedFile:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
        base_metadata = dict(shared_metadata or {})
        base_metadata["tags"] = tags or []
        documents: list[NormalizedDocument] = []

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            if not any(headers):
                continue

            for row_number, row in enumerate(rows[1:], start=2):
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue

                row_lines = []
                for column_index, header in enumerate(headers):
                    label = header or f"column_{column_index + 1}"
                    value = row[column_index] if column_index < len(row) else None
                    row_lines.append(f"{label}: {'' if value is None else value}")

                row_text = "\n".join(
                    [
                        f"Workbook: {filename}",
                        f"Sheet: {sheet.title}",
                        f"Row: {row_number}",
                        *row_lines,
                    ]
                ).strip()

                metadata = dict(base_metadata)
                metadata.update(
                    {
                        "sheet_name": sheet.title,
                        "row_start": row_number,
                        "row_end": row_number,
                        "column_headers": headers,
                    }
                )

                documents.append(
                    self.build_document(
                        title=f"{filename} {sheet.title} row {row_number}",
                        source_type=source_type_override or "xlsx",
                        content=row_text,
                        metadata=metadata,
                        original_filename=filename,
                        mime_type=mime_type,
                    )
                )

        if not documents:
            raise ValueError("xlsx file has no readable sheets or rows")

        return ParsedFile(filename=filename, detected_type="xlsx", documents=documents)
